# In this file we will create a streamlit app to generate test plans for automotive requirements with image processing capabilities that will help to extract more information from images related to the requirements and enhance the test plan generation process 
# The app will allow users to upload images such as diagrams, screenshots, or visual specifications that accompany the requirements. The AI model will analyze these images to extract relevant information that can be used to generate more comprehensive and accurate test plans.
# The app will also allow users to upload a specification document to provide additional context for the requirements.
# The app will use the OpenRouter API to access the DeepSeek R1 model for generating test plans, and it will also support local models like Ollama if available.
# The app will provide options to download the generated test plans in both Excel and JSON formats, with the Excel format matching a specific design layout.



import streamlit as st
import pandas as pd
import subprocess
import json
import io
import re
import time
from io import BytesIO
import zipfile
import os
from openai import OpenAI
import docx
from sentence_transformers import SentenceTransformer
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import base64
from PIL import Image

# Initialize session state
if 'spec_chunks' not in st.session_state:
    st.session_state.spec_chunks = []
if 'spec_embeddings' not in st.session_state:
    st.session_state.spec_embeddings = None
if 'embedding_model' not in st.session_state:
    st.session_state.embedding_model = None
if 'api_key' not in st.session_state:
    st.session_state.api_key = None

def get_api_key():
    """Get API key from session state or environment"""
    return st.session_state.api_key or os.getenv("OPENROUTER_API_KEY")

def get_openrouter_client():
    """Initialize OpenRouter client"""
    api_key = get_api_key()
    if not api_key:
        return None
    
    return OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
    )

def call_r1_distill_llama(prompt, temperature=0.7):
    """Call the DeepSeek R1 model via OpenRouter"""
    try:
        client = get_openrouter_client()
        if not client:
            st.error("OpenRouter API key not provided")
            return None
        
        response = client.chat.completions.create(
            model="deepseek/deepseek-r1-distill-llama-70b",
            messages=[{"role": "user", "content": prompt}],
            temperature=temperature,
            max_tokens=4000
        )
        
        return response.choices[0].message.content
        
    except Exception as e:
        st.error(f"OpenRouter API error: {str(e)}")
        return None

def call_ollama_model(prompt, model_name="llama3.2"):
    """Call local Ollama model"""
    try:
        import requests
        
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={
                "model": model_name,
                "prompt": prompt,
                "stream": False,
                "options": {"temperature": 0.7}
            },
            timeout=300
        )
        
        if response.status_code == 200:
            return response.json().get('response', '')
        else:
            st.error(f"Ollama error: {response.status_code}")
            return None
            
    except Exception as e:
        st.error(f"Ollama connection error: {str(e)}")
        return None

def parse_requirement(requirement_text):
    """Parse requirement into structured format"""
    parts = {
        'condition': '',
        'action': '',
        'parameters': [],
        'structure': 'unstructured'
    }
    
    # English IF-THEN pattern
    if_then_pattern = r'IF\s+(.*?)\s+THEN\s+(.*?)(?:\.|$)'
    if_then_match = re.search(if_then_pattern, requirement_text, re.IGNORECASE | re.DOTALL)
    
    # French SI-ALORS pattern
    si_alors_pattern = r'SI\s+(.*?)\s+ALORS\s+(.*?)(?:\.|$)'
    si_alors_match = re.search(si_alors_pattern, requirement_text, re.IGNORECASE | re.DOTALL)
    
    if if_then_match:
        parts['condition'] = if_then_match.group(1).strip()
        parts['action'] = if_then_match.group(2).strip()
        parts['structure'] = 'if_then'
    elif si_alors_match:
        parts['condition'] = si_alors_match.group(1).strip()
        parts['action'] = si_alors_match.group(2).strip()
        parts['structure'] = 'si_alors'
    else:
        # Try to identify condition and action patterns
        condition_keywords = ['when', 'if', 'while', 'during', 'quand', 'si', 'pendant']
        action_keywords = ['shall', 'must', 'should', 'will', 'doit', 'devra']
        
        sentences = requirement_text.split('.')
        for sentence in sentences:
            sentence = sentence.strip().lower()
            if any(keyword in sentence for keyword in condition_keywords):
                parts['condition'] = sentence
            elif any(keyword in sentence for keyword in action_keywords):
                parts['action'] = sentence
    
    return parts

def extract_parameters(requirement_text):
    """Extract parameters from requirement text"""
    parameters = []
    
    # Speed parameters
    speed_pattern = r'(\d+(?:\.\d+)?)\s*km/h'
    speed_matches = re.findall(speed_pattern, requirement_text, re.IGNORECASE)
    for speed in speed_matches:
        parameters.append(f"Speed: {speed} km/h")
    
    # Sound parameters
    sound_patterns = [r'SOUND_\w+', r'sound\s+level', r'acoustic', r'audio', r'volume']
    for pattern in sound_patterns:
        if re.search(pattern, requirement_text, re.IGNORECASE):
            parameters.append("Sound/Audio system")
            break
    
    # AVAS specific
    if re.search(r'AVAS', requirement_text, re.IGNORECASE):
        parameters.append("AVAS system")
    
    # Vehicle states
    vehicle_states = ['parked', 'moving', 'stopped', 'reverse', 'forward', 'stationary']
    for state in vehicle_states:
        if re.search(rf'\b{state}\b', requirement_text, re.IGNORECASE):
            parameters.append(f"Vehicle state: {state}")
    
    return list(set(parameters))

def generate_test_name(requirement_id, parameters):
    """Generate descriptive test name"""
    if not requirement_id:
        return "AVAS_Test_Plan"
    
    return f"{requirement_id} var0"

def extract_covered_requirements(requirement_description):
    """Extract covered requirements from description"""
    req_pattern = r'\b(REQ[-_]\w+|\w+[-_]\d+)\b'
    requirements = re.findall(req_pattern, requirement_description, re.IGNORECASE)
    return list(set(requirements)) if requirements else []

def analyze_requirement(requirement_text):
    """Analyze requirement and return structured information"""
    parsed = parse_requirement(requirement_text)
    parameters = extract_parameters(requirement_text)
    
    analysis = {
        'structure_type': parsed['structure'],
        'condition': parsed['condition'],
        'action': parsed['action'],
        'parameters': parameters,
        'complexity': 'high' if len(parameters) > 3 else 'medium' if len(parameters) > 1 else 'low',
        'test_areas': []
    }
    
    # Determine test areas
    if any('speed' in p.lower() for p in parameters):
        analysis['test_areas'].append('Speed-dependent behavior')
    if any('sound' in p.lower() or 'audio' in p.lower() for p in parameters):
        analysis['test_areas'].append('Audio/acoustic verification')
    if any('avas' in p.lower() for p in parameters):
        analysis['test_areas'].append('AVAS system functionality')
    
    return analysis

def load_embedding_model():
    """Load sentence transformer model for embeddings"""
    try:
        if st.session_state.embedding_model is None:
            with st.spinner("Loading embedding model..."):
                st.session_state.embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
        return st.session_state.embedding_model
    except Exception as e:
        st.error(f"Error loading embedding model: {str(e)}")
        return None

def extract_text_from_docx(file):
    """Extract text from DOCX file"""
    try:
        doc = docx.Document(file)
        text = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text.append(paragraph.text.strip())
        return '\n'.join(text)
    except Exception as e:
        st.error(f"Error reading DOCX file: {str(e)}")
        return ""

def chunk_text(text, chunk_size=800, overlap=200):
    """Split text into chunks with overlap"""
    if not text:
        return []
    
    chunks = []
    start = 0
    
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        
        # Try to break at sentence boundary
        if end < len(text):
            last_period = chunk.rfind('.')
            last_newline = chunk.rfind('\n')
            break_point = max(last_period, last_newline)
            
            if break_point > start + chunk_size // 2:
                chunk = text[start:break_point + 1]
                end = break_point + 1
        
        chunks.append(chunk.strip())
        start = end - overlap
        
        if start >= len(text):
            break
    
    return [chunk for chunk in chunks if len(chunk.strip()) > 50]

def create_embeddings(texts, model):
    """Create embeddings for text chunks"""
    try:
        embeddings = model.encode(texts)
        return embeddings
    except Exception as e:
        st.error(f"Error creating embeddings: {str(e)}")
        return None

def search_specification(query, chunks, embeddings, model, top_k=5):
    """Search specification chunks using semantic similarity"""
    try:
        query_embedding = model.encode([query])
        similarities = cosine_similarity(query_embedding, embeddings)[0]
        
        # Get top-k most similar chunks
        top_indices = np.argsort(similarities)[-top_k:][::-1]
        results = [(chunks[i], similarities[i]) for i in top_indices if similarities[i] > 0.3]
        
        return results
    except Exception as e:
        st.error(f"Error searching specification: {str(e)}")
        return []

def process_specification_file(uploaded_file):
    """Process uploaded specification file"""
    try:
        # Extract text
        if uploaded_file.type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            text = extract_text_from_docx(uploaded_file)
        else:
            text = str(uploaded_file.read(), "utf-8")
        
        if not text:
            st.error("Could not extract text from the file")
            return
        
        # Load embedding model
        model = load_embedding_model()
        if not model:
            return
        
        # Create chunks
        with st.spinner("Processing specification..."):
            chunks = chunk_text(text)
            
            if not chunks:
                st.error("No valid chunks created from specification")
                return
            
            # Create embeddings
            embeddings = create_embeddings(chunks, model)
            
            if embeddings is not None:
                st.session_state.spec_chunks = chunks
                st.session_state.spec_embeddings = embeddings
                st.success(f"✅ Specification processed: {len(chunks)} chunks created")
            else:
                st.error("Failed to create embeddings")
                
    except Exception as e:
        st.error(f"Error processing specification: {str(e)}")

def process_image_with_ai(image_file, requirement_description):
    """Process uploaded image to extract relevant information for test plan generation"""
    try:
        # Convert image to base64 for API
        image = Image.open(image_file)
        buffered = BytesIO()
        image.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode()
        
        client = get_openrouter_client()
        
        response = client.chat.completions.create(
            model="openai/gpt-4o",  # Use vision-capable model
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": f"""Analyze this image in the context of automotive AVAS (Acoustic Vehicle Alerting System) testing.

Requirement Description: {requirement_description}

Please extract and describe:
1. Any visual specifications, parameters, or conditions shown
2. UI elements, diagrams, or system states visible
3. Relevant technical details that would impact test plan generation
4. Any specific values, thresholds, or operational modes depicted

Provide a detailed description that can be used as additional context for generating automotive test plans."""
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/png;base64,{img_str}"
                            }
                        }
                    ]
                }
            ],
            max_tokens=1000
        )
        
        return response.choices[0].message.content
        
    except Exception as e:
        st.error(f"Error processing image: {str(e)}")
        return None

def generate_test_plan_with_image_context(requirement_id, req_description, image_context=""):
    """Generate test plan with additional image context"""
    try:
        # Parse requirement
        req_parts = parse_requirement(req_description)
        parameters = extract_parameters(req_description)
        
        # Search specification if available
        spec_context = ""
        spec_references = []
        if st.session_state.spec_chunks and st.session_state.spec_embeddings is not None:
            search_results = search_specification(
                req_description,
                st.session_state.spec_chunks,
                st.session_state.spec_embeddings,
                st.session_state.embedding_model
            )
            
            if search_results:
                spec_context = "\n".join([f"Spec: {chunk}" for chunk, _ in search_results[:3]])
                spec_references = [
                    {
                        "parameter": param,
                        "spec_reference": search_results[0][0] if search_results else "",
                        "relevance_score": search_results[0][1] if search_results else 0,
                        "explanation": f"Related to {param} in requirement"
                    } for param in parameters
                ]
        
        # Enhanced prompt with image context
        prompt = f"""You are an expert automotive AVAS (Acoustic Vehicle Alerting System) test engineer.

REQUIREMENT ID: {requirement_id}
REQUIREMENT: {req_description}

PARSED REQUIREMENT STRUCTURE:
- Condition: {req_parts.get('condition', 'N/A')}
- Action: {req_parts.get('action', 'N/A')}
- Parameters: {parameters}

{f"VISUAL CONTEXT FROM IMAGE: {image_context}" if image_context else ""}

{f"SPECIFICATION CONTEXT: {spec_context}" if spec_context else ""}

Generate a comprehensive test plan with the following structure:

1. TEST IDENTIFICATION:
   - Test Name: Should be "{requirement_id} var0" format
   - Variant: Always "var0" 
   - Description: Clear test objective
   - Covered Requirements: {requirement_id}

2. TEST STEPS with three types:
   - CI (Initial Conditions): Setup, preconditions, system initialization
   - AC (Actions): Test execution, stimulations, commands
   - RA (Results): Verification, validation, expected results

Consider the visual context provided and ensure test steps align with any diagrams, UI elements, or specifications shown in the image.

Return JSON format:
{{
    "test_name": "{requirement_id} var0",
    "variant": "var0",
    "requirement_id": "{requirement_id}",
    "description": "test objective",
    "covered_requirement": "{requirement_id}",
    "steps": [
        {{"step_number": "1", "type": "CI", "description": "step details"}}
    ]
}}"""

        response = call_r1_distill_llama(prompt, temperature=0.3)
        
        if response:
            # Parse JSON response
            json_match = re.search(r'\{.*\}', response, re.DOTALL)
            if json_match:
                test_plan = json.loads(json_match.group())
                
                # Add specification references if available
                if spec_references:
                    test_plan['spec_references'] = spec_references
                
                # Add image context to test plan
                if image_context:
                    test_plan['image_context'] = image_context
                
                return test_plan
        
        return None
        
    except Exception as e:
        st.error(f"Error generating test plan: {str(e)}")
        return None

def create_enhanced_excel_file(test_plan):
    """Create Excel file matching the specified design layout"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create the main test plan sheet with the specified layout
        create_main_test_sheet(writer, test_plan)
        
        # Add specification references sheet if available
        if 'spec_references' in test_plan:
            create_spec_references_sheet(writer, test_plan)
    
    output.seek(0)
    return output.getvalue()

def create_main_test_sheet(writer, test_plan):
    """Create the main test sheet with the specified layout"""
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    
    # Prepare data in the exact format shown in the image
    test_data = []
    
    for i, step in enumerate(test_plan.get('steps', [])):
        row_data = {
            'Test Name': test_plan.get('test_name', ''),
            'Requirement': test_plan.get('requirement_id', ''),
            'Version': test_plan.get('variant', 'var0'),
            'Test Description': test_plan.get('description', ''),
            'Step Type': step.get('type', ''),
            'Step n°': step.get('step_number', str(i+1)),
            'Step Description': step.get('description', ''),
            'Tested Requirements': test_plan.get('covered_requirement', '')
        }
        test_data.append(row_data)
    
    # Create DataFrame
    df = pd.DataFrame(test_data)
    
    # Write to Excel
    df.to_excel(writer, sheet_name='Test Plan', index=False, startrow=1)
    
    # Get the workbook and worksheet to apply formatting
    workbook = writer.book
    worksheet = writer.sheets['Test Plan']
    
    # Apply the enhanced formatting
    apply_enhanced_formatting(worksheet, test_plan, len(test_data))

def apply_enhanced_formatting(worksheet, test_plan, num_rows):
    """Apply enhanced formatting to match the design"""
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    
    # Define colors and styles
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ci_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")  # Light blue
    ac_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    ra_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Light green
    
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    header_font = Font(bold=True, size=10)
    cell_font = Font(size=9)
    
    # Set column widths
    column_widths = {
        'A': 15,  # Test Name
        'B': 15,  # Requirement
        'C': 10,  # Version
        'D': 30,  # Test Description
        'E': 8,   # Step Type
        'F': 8,   # Step n°
        'G': 50,  # Step Description
        'H': 20   # Tested Requirements
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # Format headers (row 2, since we started at row 1)
    for col in range(1, 9):
        cell = worksheet.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Format data rows
    for row in range(3, 3 + num_rows):
        step_type_cell = worksheet.cell(row=row, column=5)  # Step Type column
        step_type = step_type_cell.value
        
        # Apply row formatting based on step type
        for col in range(1, 9):
            cell = worksheet.cell(row=row, column=col)
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Color code based on step type
            if step_type == 'CI':
                cell.fill = ci_fill
            elif step_type == 'AC':
                cell.fill = ac_fill
            elif step_type == 'RA':
                cell.fill = ra_fill
    
    # Add title row
    worksheet.insert_rows(1)
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = f"Test Plan: {test_plan.get('test_name', '')}"
    title_cell.font = Font(bold=True, size=12)
    worksheet.merge_cells('A1:H1')

def create_spec_references_sheet(writer, test_plan):
    """Create specification references sheet"""
    if 'spec_references' not in test_plan:
        return
    
    spec_data = []
    for ref in test_plan['spec_references']:
        spec_data.append({
            'Parameter': ref.get('parameter', ''),
            'Specification Reference': ref.get('spec_reference', ''),
            'Relevance Score': ref.get('relevance_score', 0),
            'Explanation': ref.get('explanation', '')
        })
    
    df_spec = pd.DataFrame(spec_data)
    df_spec.to_excel(writer, sheet_name='Specification References', index=False)

def display_enhanced_test_plan(test_plan):
    """Display enhanced test plan with image context"""
    
    # Main test information in a more structured way
    st.markdown("### 📋 Test Plan Information")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown(f"""
        **Test Name:** {test_plan.get('test_name', 'N/A')}  
        **Variant:** {test_plan.get('variant', 'var0')}  
        **Requirement ID:** {test_plan.get('requirement_id', 'N/A')}
        """)
    
    with col2:
        st.markdown(f"""
        **Description:** {test_plan.get('description', 'N/A')}  
        **Covered Requirements:** {test_plan.get('covered_requirement', 'N/A')}
        """)
    
    # Display image context if available
    if 'image_context' in test_plan and test_plan['image_context']:
        st.markdown("### 📷 Visual Context Analysis")
        st.info(test_plan['image_context'])
    
    # Test Steps with enhanced formatting
    st.markdown("### 🔄 Test Steps")
    
    steps_data = []
    for step in test_plan.get('steps', []):
        steps_data.append({
            "Step N°": step.get('step_number', ''),
            "Type": step.get('type', ''),
            "Description": step.get('description', ''),
        })
    
    if steps_data:
        df = pd.DataFrame(steps_data)
        
        # Enhanced styling function
        def style_steps(styler):
            def color_step_type(val):
                if val == 'CI':
                    return 'background-color: #B7DEE8; color: black'
                elif val == 'AC':
                    return 'background-color: #FFFF00; color: black'
                elif val == 'RA':
                    return 'background-color: #92D050; color: black'
                return ''
            
            return styler.applymap(color_step_type, subset=['Type'])
        
        styled_df = df.style.pipe(style_steps)
        st.dataframe(styled_df, use_container_width=True, hide_index=True)
    
    # Specification references
    if 'spec_references' in test_plan and test_plan['spec_references']:
        st.markdown("### 📖 Specification References")
        for i, ref in enumerate(test_plan['spec_references']):
            with st.expander(f"Reference {i+1}: {ref.get('parameter', 'General')}"):
                st.write(f"**Relevance Score:** {ref.get('relevance_score', 0):.2f}")
                st.write(f"**Context:** {ref.get('explanation', 'N/A')}")
                st.code(ref.get('spec_reference', 'N/A'))

def enhanced_single_test_tab():
    st.header("🧪 Single Test Plan Generator")
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.subheader("Requirement Input")
        requirement_id = st.text_input("Requirement ID:", placeholder="REQ-AVAS-001")
        requirement_description = st.text_area(
            "Requirement Description:", 
            placeholder="IF vehicle speed < 20 km/h THEN AVAS shall emit sound",
            height=150
        )
        
        # Image upload section
        st.subheader("📷 Visual Context (Optional)")
        uploaded_image = st.file_uploader(
            "Upload requirement diagram/screenshot",
            type=['png', 'jpg', 'jpeg'],
            help="Upload screenshots, diagrams, or visual specifications that accompany this requirement"
        )
        
        # Display uploaded image
        if uploaded_image:
            st.image(uploaded_image, caption="Uploaded Requirement Visual", use_column_width=True)
    
    with col2:
        st.subheader("Requirement Analysis")
        if requirement_description:
            analysis = analyze_requirement(requirement_description)
            st.json(analysis)
        
        # Process image if uploaded
        image_context = ""
        if uploaded_image and st.button("🔍 Analyze Image"):
            with st.spinner("Processing image..."):
                image_analysis = process_image_with_ai(uploaded_image, requirement_description)
                if image_analysis:
                    st.subheader("Image Analysis")
                    st.write(image_analysis)
                    image_context = image_analysis
    
    # Generate button
    if st.button("🚀 Generate Enhanced Test Plan", type="primary"):
        if requirement_id and requirement_description:
            with st.spinner("Generating test plan with visual context..."):
                # Enhanced generation with image context
                test_plan = generate_test_plan_with_image_context(
                    requirement_id, 
                    requirement_description, 
                    image_context
                )
                
                if test_plan:
                    st.success("✅ Test plan generated successfully!")
                    display_enhanced_test_plan(test_plan)
                    
                    # Download options
                    col1, col2 = st.columns(2)
                    with col1:
                        excel_data = create_enhanced_excel_file(test_plan)
                        st.download_button(
                            "📊 Download Excel",
                            data=excel_data,
                            file_name=f"test_plan_{requirement_id}_{test_plan.get('variant', 'var0')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    with col2:
                        json_data = json.dumps(test_plan, indent=2, ensure_ascii=False)
                        st.download_button(
                            "📄 Download JSON",
                            data=json_data,
                            file_name=f"test_plan_{requirement_id}.json",
                            mime="application/json"
                        )
        else:
            st.warning("Please provide both Requirement ID and Description.")

def generate_single_test_plan(requirement_id, req_description):
    """Generate test plan for single requirement (fallback function)"""
    try:
        req_parts = parse_requirement(req_description)
        parameters = extract_parameters(req_description)
        
        spec_context = ""
        spec_references = []
        if st.session_state.spec_chunks and st.session_state.spec_embeddings is not None:
            search_results = search_specification(
                req_description,
                st.session_state.spec_chunks,
                st.session_state.spec_embeddings,
                st.session_state.embedding_model
            )
            
            if search_results:
                spec_context = "\n".join([f"Spec: {chunk}" for chunk, _ in search_results[:3]])
                spec_references = [
                    {
                        "parameter": param,
                        "spec_reference": search_results[0][0] if search_results else "",
                        "relevance_score": search_results[0][1] if search_results else 0,
                        "explanation": f"Related to {param} in requirement"
                    } for param in parameters
                ]

        prompt = f"""You are an expert automotive AVAS (Acoustic Vehicle Alerting System) test engineer.

REQUIREMENT ID: {requirement_id}
REQUIREMENT: {req_description}

PARSED REQUIREMENT STRUCTURE:
- Condition: {req_parts.get('condition', 'N/A')}
- Action: {req_parts.get('action', 'N/A')}
- Parameters: {parameters}

{f"SPECIFICATION CONTEXT: {spec_context}" if spec_context else ""}

Generate a comprehensive test plan for this AVAS requirement with the following structure:

1. TEST IDENTIFICATION:
   - Test Name: Should be "{requirement_id} var0" format
   - Variant: Always "var0"
   - Description: Clear description of what this test validates
   - Covered Requirements: {requirement_id}

2. TEST STEPS with three types:
   - CI (Initial Conditions): System setup, preconditions, initialization steps
   - AC (Actions): Test execution steps, stimulations, commands to execute
   - RA (Results): Verification steps, expected results, validation criteria

Focus on AVAS-specific testing including:
- Vehicle speed conditions
- Sound emission verification  
- System state validation
- Acoustic measurements

Return ONLY a valid JSON object with this exact structure:
{{
    "test_name": "{requirement_id} var0",
    "variant": "var0", 
    "requirement_id": "{requirement_id}",
    "description": "Test description here",
    "covered_requirement": "{requirement_id}",
    "steps": [
        {{"step_number": "1", "type": "CI", "description": "Initial condition description"}},
        {{"step_number": "2", "type": "AC", "description": "Action description"}},
        {{"step_number": "3", "type": "RA", "description": "Result verification description"}}
    ]
}}"""

        response = call_r1_distill_llama(prompt, temperature=0.3)
        
        if response:
            json_match = re.search(r'\{.*\}', response, re.DOTALL)
            if json_match:
                test_plan = json.loads(json_match.group())
                
                if spec_references:
                    test_plan['spec_references'] = spec_references
                
                return test_plan
        
        return None
        
    except Exception as e:
        st.error(f"Error generating test plan: {str(e)}")
        return None

def batch_processing_tab():
    """Batch processing tab for multiple requirements"""
    st.header("📊 Batch Test Plan Generation")
    
    st.markdown("""
    Upload a CSV or Excel file containing multiple requirements for batch processing.
    
    **Required columns:**
    - `requirement_id`: Unique requirement identifier
    - `requirement_description`: Full requirement text
    """)
    
    # File upload
    uploaded_file = st.file_uploader(
        "Upload Requirements File",
        type=['csv', 'xlsx', 'xls'],
        help="CSV or Excel file with requirement_id and requirement_description columns"
    )
    
    if uploaded_file:
        try:
            # Read file
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            
            st.subheader("📋 File Preview")
            st.dataframe(df.head(), use_container_width=True)
            
            # Validate required columns
            required_cols = ['requirement_id', 'requirement_description']
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                st.error(f"Missing required columns: {missing_cols}")
                st.info(f"Available columns: {list(df.columns)}")
            else:
                st.success(f"✅ Valid file format. Found {len(df)} requirements.")
                
                # Processing options
                col1, col2 = st.columns(2)
                with col1:
                    max_requirements = st.number_input(
                        "Max requirements to process:", 
                        min_value=1, 
                        max_value=len(df), 
                        value=min(10, len(df))
                    )
                
                with col2:
                    include_images = st.checkbox("Include image analysis (slower)", value=False)
                
                if st.button("🚀 Generate Batch Test Plans", type="primary"):
                    generate_batch_test_plans(df, max_requirements, include_images)
                    
        except Exception as e:
            st.error(f"Error reading file: {str(e)}")

def generate_batch_test_plans(df, max_requirements, include_images=False):
    """Generate test plans for multiple requirements"""
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    all_test_plans = []
    failed_requirements = []
    
    # Process requirements
    for i, row in df.head(max_requirements).iterrows():
        try:
            progress = (i + 1) / max_requirements
            progress_bar.progress(progress)
            status_text.text(f"Processing requirement {i+1}/{max_requirements}: {row['requirement_id']}")
            
            # Generate test plan
            if include_images:
                # For batch processing, we can't process images individually
                # This would require additional image columns in the CSV
                test_plan = generate_single_test_plan(
                    row['requirement_id'], 
                    row['requirement_description']
                )
            else:
                test_plan = generate_single_test_plan(
                    row['requirement_id'], 
                    row['requirement_description']
                )
            
            if test_plan:
                all_test_plans.append(test_plan)
            else:
                failed_requirements.append(row['requirement_id'])
                
            # Small delay to prevent API rate limiting
            time.sleep(1)
            
        except Exception as e:
            st.error(f"Error processing {row['requirement_id']}: {str(e)}")
            failed_requirements.append(row['requirement_id'])
    
    # Results summary
    progress_bar.progress(1.0)
    status_text.text("Processing complete!")
    
    st.subheader("📈 Batch Processing Results")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Processed", len(all_test_plans) + len(failed_requirements))
    with col2:
        st.metric("Successful", len(all_test_plans))
    with col3:
        st.metric("Failed", len(failed_requirements))
    
    if failed_requirements:
        st.warning(f"Failed to process: {', '.join(failed_requirements)}")
    
    # Download options
    if all_test_plans:
        st.subheader("💾 Download Results")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Create combined Excel file
            excel_data = create_batch_excel_file(all_test_plans)
            st.download_button(
                "📊 Download Combined Excel",
                data=excel_data,
                file_name="batch_test_plans.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            # Create ZIP archive
            zip_data = create_zip_archive(all_test_plans)
            st.download_button(
                "📦 Download ZIP Archive", 
                data=zip_data,
                file_name="test_plans_archive.zip",
                mime="application/zip"
            )

def create_batch_excel_file(test_plans):
    """Create combined Excel file for batch test plans"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create summary sheet
        create_batch_summary_sheet(writer, test_plans)
        
        # Create individual test plan sheets
        for i, test_plan in enumerate(test_plans[:10]):  # Limit to 10 sheets
            sheet_name = f"Test_{test_plan.get('requirement_id', f'Plan_{i+1}')}"[:31]
            create_main_test_sheet(writer, test_plan)
            
        # Create combined data sheet
        create_combined_data_sheet(writer, test_plans)
    
    output.seek(0)
    return output.getvalue()

def create_batch_summary_sheet(writer, test_plans):
    """Create summary sheet for batch processing"""
    summary_data = []
    
    for test_plan in test_plans:
        summary_data.append({
            'Test Name': test_plan.get('test_name', ''),
            'Requirement ID': test_plan.get('requirement_id', ''),
            'Variant': test_plan.get('variant', 'var0'),
            'Description': test_plan.get('description', ''),
            'Total Steps': len(test_plan.get('steps', [])),
            'CI Steps': len([s for s in test_plan.get('steps', []) if s.get('type') == 'CI']),
            'AC Steps': len([s for s in test_plan.get('steps', []) if s.get('type') == 'AC']),
            'RA Steps': len([s for s in test_plan.get('steps', []) if s.get('type') == 'RA']),
            'Has Specifications': 'Yes' if 'spec_references' in test_plan else 'No'
        })
    
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Summary', index=False)

def create_combined_data_sheet(writer, test_plans):
    """Create combined data sheet with all test plans"""
    all_steps = []
    
    for test_plan in test_plans:
        for step in test_plan.get('steps', []):
            all_steps.append({
                'Test Name': test_plan.get('test_name', ''),
                'Requirement': test_plan.get('requirement_id', ''),
                'Version': test_plan.get('variant', 'var0'),
                'Test Description': test_plan.get('description', ''),
                'Step Type': step.get('type', ''),
                'Step n°': step.get('step_number', ''),
                'Step Description': step.get('description', ''),
                'Tested Requirements': test_plan.get('covered_requirement', '')
            })
    
    df_combined = pd.DataFrame(all_steps)
    df_combined.to_excel(writer, sheet_name='All Test Steps', index=False)

def create_zip_archive(test_plans):
    """Create ZIP archive with individual files"""
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Add individual Excel files
        for test_plan in test_plans:
            excel_data = create_enhanced_excel_file(test_plan)
            filename = f"test_plan_{test_plan.get('requirement_id', 'unknown')}.xlsx"
            zip_file.writestr(filename, excel_data)
        
        # Add combined JSON file
        combined_json = json.dumps(test_plans, indent=2, ensure_ascii=False)
        zip_file.writestr("all_test_plans.json", combined_json)
        
        # Add summary report
        summary_text = generate_summary_report(test_plans)
        zip_file.writestr("summary_report.txt", summary_text)
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

def generate_summary_report(test_plans):
    """Generate text summary report"""
    report = f"""Test Plans Generation Summary Report
Generated on: {time.strftime('%Y-%m-%d %H:%M:%S')}

OVERVIEW:
- Total test plans generated: {len(test_plans)}
- Total test steps: {sum(len(plan.get('steps', [])) for plan in test_plans)}

STEP TYPE BREAKDOWN:
- CI (Initial Conditions): {sum(len([s for s in plan.get('steps', []) if s.get('type') == 'CI']) for plan in test_plans)}
- AC (Actions): {sum(len([s for s in plan.get('steps', []) if s.get('type') == 'AC']) for plan in test_plans)}
- RA (Results): {sum(len([s for s in plan.get('steps', []) if s.get('type') == 'RA']) for plan in test_plans)}

TEST PLANS GENERATED:
"""
    
    for i, plan in enumerate(test_plans, 1):
        report += f"{i}. {plan.get('test_name', 'Unknown')} - {plan.get('description', 'No description')}\n"
    
    return report

def analysis_insights_tab():
    """Analysis and insights tab"""
    st.header("📈 Analysis & Insights")
    
    if not hasattr(st.session_state, 'generated_test_plans') or not st.session_state.generated_test_plans:
        st.info("Generate some test plans first to see analysis and insights.")
        return
    
    test_plans = st.session_state.generated_test_plans
    
    # Overview metrics
    st.subheader("📊 Overview Metrics")
    
    col1, col2, col3, col4 = st.columns(4)
    
    total_plans = len(test_plans)
    total_steps = sum(len(plan.get('steps', [])) for plan in test_plans)
    avg_steps = total_steps / total_plans if total_plans > 0 else 0
    
    with col1:
        st.metric("Total Test Plans", total_plans)
    with col2:
        st.metric("Total Test Steps", total_steps)
    with col3:
        st.metric("Avg Steps per Plan", f"{avg_steps:.1f}")
    with col4:
        complexity_score = calculate_complexity_score(test_plans)
        st.metric("Complexity Score", f"{complexity_score:.1f}")
    
    # Step type analysis
    st.subheader("🔄 Step Type Distribution")
    
    step_counts = {'CI': 0, 'AC': 0, 'RA': 0}
    for plan in test_plans:
        for step in plan.get('steps', []):
            step_type = step.get('type', 'Unknown')
            if step_type in step_counts:
                step_counts[step_type] += 1
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Bar chart
        df_steps = pd.DataFrame(list(step_counts.items()), columns=['Step Type', 'Count'])
        st.bar_chart(df_steps.set_index('Step Type'))
    
    with col2:
        # Metrics
        for step_type, count in step_counts.items():
            percentage = (count / total_steps * 100) if total_steps > 0 else 0
            st.metric(f"{step_type} Steps", f"{count} ({percentage:.1f}%)")
    
    # Requirement analysis
    st.subheader("📋 Requirement Analysis")
    
    requirement_stats = analyze_requirements_patterns(test_plans)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Common Patterns:**")
        for pattern, count in requirement_stats['patterns'].items():
            st.write(f"- {pattern}: {count} occurrences")
    
    with col2:
        st.write("**Parameter Distribution:**")
        for param, count in requirement_stats['parameters'].items():
            st.write(f"- {param}: {count} requirements")

def calculate_complexity_score(test_plans):
    """Calculate complexity score for test plans"""
    if not test_plans:
        return 0
    
    total_score = 0
    for plan in test_plans:
        steps = plan.get('steps', [])
        step_count = len(steps)
        
        # Base score from step count
        score = step_count * 2
        
        # Bonus for step variety
        step_types = set(step.get('type') for step in steps)
        score += len(step_types) * 5
        
        # Bonus for specification references
        if 'spec_references' in plan:
            score += len(plan['spec_references']) * 3
        
        total_score += score
    
    return total_score / len(test_plans)

def analyze_requirements_patterns(test_plans):
    """Analyze patterns in requirements"""
    patterns = {}
    parameters = {}
    
    for plan in test_plans:
        req_id = plan.get('requirement_id', '')
        description = plan.get('description', '')
        
        # Analyze requirement ID patterns
        if req_id:
            pattern = re.match(r'([A-Z]+)', req_id)
            if pattern:
                prefix = pattern.group(1)
                patterns[prefix] = patterns.get(prefix, 0) + 1
        
        # Analyze description for common parameters
        desc_lower = description.lower()
        param_keywords = ['speed', 'sound', 'avas', 'vehicle', 'system', 'audio']
        for keyword in param_keywords:
            if keyword in desc_lower:
                parameters[keyword] = parameters.get(keyword, 0) + 1
    
    return {
        'patterns': dict(sorted(patterns.items(), key=lambda x: x[1], reverse=True)[:5]),
        'parameters': dict(sorted(parameters.items(), key=lambda x: x[1], reverse=True)[:5])
    }

def main():
    """Main application function"""
    st.set_page_config(
        page_title="🚗 Enhanced Automotive Test Plan Generator",
        page_icon="🧪",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Custom CSS for better styling
    st.markdown("""
    <style>
    .main > div {
        padding-top: 2rem;
    }
    .stDownloadButton button {
        width: 100%;
    }
    .metric-container {
        padding: 1rem;
        border-radius: 0.5rem;
        border: 1px solid #e0e0e0;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.title("🚗 Enhanced Automotive Test Plan Generator")
    st.markdown("**AI-Powered AVAS Test Plan Generation with Visual Context and Specification Integration**")
    
    # Initialize session state for generated plans
    if 'generated_test_plans' not in st.session_state:
        st.session_state.generated_test_plans = []
    
    # API Key input (if needed)
    if not get_api_key():
        with st.sidebar:
            st.header("🔑 Configuration")
            st.session_state.api_key = st.text_input(
                "OpenRouter API Key:", 
                type="password",
                help="Get your key from https://openrouter.ai"
            )
            
            if st.session_state.api_key:
                st.success("✅ API key configured")
    
    # Specification upload section
    with st.sidebar:
        st.header("📄 Specification Documents")
        uploaded_spec = st.file_uploader(
            "Upload specification (optional)",
            type=['docx', 'txt'],
            help="Upload AVAS specification for enhanced test generation"
        )
        
        if uploaded_spec and uploaded_spec != st.session_state.get('current_spec_file'):
            process_specification_file(uploaded_spec)
            st.session_state.current_spec_file = uploaded_spec
        
        # Specification search
        if st.session_state.spec_chunks:
            st.subheader("🔍 Search Specification")
            search_query = st.text_input("Search terms:")
            if search_query:
                search_results = search_specification(
                    search_query,
                    st.session_state.spec_chunks,
                    st.session_state.spec_embeddings,
                    st.session_state.embedding_model,
                    top_k=3
                )
                for i, (chunk, score) in enumerate(search_results):
                    with st.expander(f"Result {i+1} (Score: {score:.2f})"):
                        st.text(chunk)
    
    # Main tabs
    tab1, tab2, tab3 = st.tabs(["🧪 Single Test Plan", "📊 Batch Processing", "📈 Analysis & Insights"])
    
    with tab1:
        enhanced_single_test_tab()
    
    with tab2:
        batch_processing_tab()
    
    with tab3:
        analysis_insights_tab()
    
    # Footer
    st.markdown("---")
    st.markdown(
        "Created by imad for Automotive Testing | "
        "Supports AVAS requirements | "
        "Powered by AI"
    )

if __name__ == "__main__":
    main()

